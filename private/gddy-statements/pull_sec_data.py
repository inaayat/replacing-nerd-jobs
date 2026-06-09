#!/usr/bin/env python3
"""Pull GoDaddy financial data from SEC EDGAR XBRL API and build Excel workbook.

Source: https://data.sec.gov/api/xbrl/companyfacts/CIK0001609711.json
All figures directly from XBRL filings — no manual transcription.
"""

import json
import urllib.request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Load data (pre-downloaded via curl from SEC API) ──
import pathlib
data_path = pathlib.Path(__file__).parent / "sec_data.json"
with open(data_path) as f:
    raw = json.load(f)

facts = raw["facts"]["us-gaap"]

# ── Helper: extract annual 10-K values for a concept ──
FISCAL_YEARS = [2020, 2021, 2022, 2023, 2024, 2025]

def get_annual(concept, unit="USD", duration=True):
    """Get annual values for a concept using end-date to identify fiscal year.
    GoDaddy FY ends Dec 31, so end='2025-12-31' = FY2025."""
    if concept not in facts:
        return {y: None for y in FISCAL_YEARS}

    entries = facts[concept]["units"].get(unit, [])
    results = {}

    for entry in entries:
        form = entry.get("form", "")
        if form != "10-K":
            continue
        fp = entry.get("fp", "")
        if fp != "FY":
            continue

        end = entry.get("end", "")
        if not end:
            continue

        # Determine fiscal year from end date (GDDY FY ends Dec 31)
        end_year = int(end[:4])
        end_month = int(end[5:7])
        fy = end_year if end_month >= 10 else end_year - 1

        # For duration items (P&L, CF), require ~12 month span
        if duration:
            start = entry.get("start", "")
            if start:
                start_year = int(start[:4])
                # Must span roughly a full year
                if end_year - start_year == 0 and end_month - int(start[5:7]) < 10:
                    continue

        if fy in FISCAL_YEARS:
            val = entry.get("val")
            # Prefer the latest filing's value (later entries override)
            results[fy] = val

    return {y: results.get(y) for y in FISCAL_YEARS}


def millions(values):
    """Convert from raw USD to millions."""
    return {y: round(v / 1_000_000, 1) if v is not None else None for y, v in values.items()}


def per_share(concept):
    """Get per-share values."""
    return get_annual(concept, unit="USD/shares")


# ── Pull all data ──
print("Pulling data from SEC EDGAR API...")

revenue = millions(get_annual("RevenueFromContractWithCustomerExcludingAssessedTax"))
cost_rev = millions(get_annual("CostOfGoodsAndServicesSold"))
if all(v is None for v in cost_rev.values()):
    cost_rev = millions(get_annual("CostOfRevenue"))

gross_profit = millions(get_annual("GrossProfit"))
op_income = millions(get_annual("OperatingIncomeLoss"))
net_income = millions(get_annual("NetIncomeLoss"))
interest_exp = millions(get_annual("InterestExpense"))
income_tax = millions(get_annual("IncomeTaxExpenseBenefit"))
pretax_income = millions(get_annual("IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest"))

# Operating expenses breakdown
rd_expense = millions(get_annual("ResearchAndDevelopmentExpense"))
selling_expense = millions(get_annual("SellingAndMarketingExpense"))
ga_expense = millions(get_annual("GeneralAndAdministrativeExpense"))
sbc = millions(get_annual("ShareBasedCompensation"))

# EPS
eps_diluted = get_annual("EarningsPerShareDiluted", unit="USD/shares")
shares_diluted = millions(get_annual("WeightedAverageNumberOfDilutedSharesOutstanding", unit="shares"))

# Balance Sheet (point in time)
total_assets = millions(get_annual("Assets"))
current_assets = millions(get_annual("AssetsCurrent"))
cash = millions(get_annual("CashAndCashEquivalentsAtCarryingValue"))
receivables = millions(get_annual("AccountsReceivableNetCurrent"))
total_liabilities = millions(get_annual("Liabilities"))
current_liabilities = millions(get_annual("LiabilitiesCurrent"))
equity = millions(get_annual("StockholdersEquity"))
goodwill = millions(get_annual("Goodwill"))
intangibles = millions(get_annual("IntangibleAssetsNetExcludingGoodwill"))
lt_debt = millions(get_annual("LongTermDebtNoncurrent"))
if all(v is None for v in lt_debt.values()):
    lt_debt = millions(get_annual("LongTermDebt"))
deferred_rev_current = millions(get_annual("ContractWithCustomerLiabilityCurrent"))
deferred_rev_noncurrent = millions(get_annual("ContractWithCustomerLiabilityNoncurrent"))
ppe = millions(get_annual("PropertyPlantAndEquipmentNet"))

# Cash Flow
ocf = millions(get_annual("NetCashProvidedByUsedInOperatingActivities"))
capex = millions(get_annual("PaymentsToAcquirePropertyPlantAndEquipment"))
acquisitions = millions(get_annual("PaymentsToAcquireBusinessesNetOfCashAcquired"))
investing = millions(get_annual("NetCashProvidedByUsedInInvestingActivities"))
financing = millions(get_annual("NetCashProvidedByUsedInFinancingActivities"))
repurchases = millions(get_annual("PaymentsForRepurchaseOfCommonStock"))
da = millions(get_annual("DepreciationAndAmortization"))
if all(v is None for v in da.values()):
    da = millions(get_annual("DepreciationDepletionAndAmortization"))

print("Data pulled successfully. Building Excel...")

# ── Build Excel ──
wb = Workbook()

header_font = Font(name="Calibri", bold=True, size=11)
title_font = Font(name="Calibri", bold=True, size=14)
section_font = Font(name="Calibri", bold=True, size=11, color="2C2E25")
header_fill = PatternFill(start_color="D4E4C8", end_color="D4E4C8", fill_type="solid")
section_fill = PatternFill(start_color="F0E8C0", end_color="F0E8C0", fill_type="solid")
thin_border = Border(bottom=Side(style='thin', color='CCCCCC'))
number_fmt = '#,##0.0'
pct_fmt = '0.0%'


def write_sheet(ws, title, subtitle, sections):
    """Write a financial statement sheet."""
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A2'] = subtitle
    ws['A2'].font = Font(name="Calibri", size=9, italic=True, color="595B4A")

    # Headers
    headers = [""] + [f"FY{y}" for y in FISCAL_YEARS]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')

    row = 5
    for section_name, items in sections:
        # Section header
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = section_fill
        ws.cell(row=row, column=1, value=section_name).font = section_font
        row += 1

        for label, data in items:
            ws.cell(row=row, column=1, value=label).border = thin_border
            for i, y in enumerate(FISCAL_YEARS, 2):
                cell = ws.cell(row=row, column=i)
                val = data.get(y) if isinstance(data, dict) else None
                cell.value = val
                if val is not None:
                    if isinstance(val, float) and abs(val) < 1:
                        cell.number_format = pct_fmt
                    else:
                        cell.number_format = number_fmt
                cell.alignment = Alignment(horizontal='right')
                cell.border = thin_border
            row += 1

    ws.column_dimensions['A'].width = 38
    from openpyxl.utils import get_column_letter
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


# ── Income Statement ──
ws_is = wb.active
ws_is.title = "Income Statement"

# Compute derived values
total_opex = {}
gross_margin = {}
op_margin = {}
net_margin = {}
for y in FISCAL_YEARS:
    r = revenue.get(y)
    gp = gross_profit.get(y)
    oi = op_income.get(y)
    ni = net_income.get(y)
    gross_margin[y] = round(gp / r, 3) if (gp and r) else None
    op_margin[y] = round(oi / r, 3) if (oi and r) else None
    net_margin[y] = round(ni / r, 3) if (ni and r) else None

write_sheet(ws_is, "Income Statement", "GoDaddy Inc. (GDDY) — Source: SEC EDGAR XBRL — All figures in $M", [
    ("Revenue", [
        ("Total Revenue", revenue),
    ]),
    ("Cost & Gross Profit", [
        ("Cost of Revenue", cost_rev),
        ("Gross Profit", gross_profit),
        ("Gross Margin", gross_margin),
    ]),
    ("Operating Expenses", [
        ("Research & Development", rd_expense),
        ("Selling & Marketing", selling_expense),
        ("General & Administrative", ga_expense),
        ("Stock-Based Compensation", sbc),
    ]),
    ("Operating & Net Income", [
        ("Operating Income (EBIT)", op_income),
        ("Operating Margin", op_margin),
        ("Interest Expense", interest_exp),
        ("Pre-Tax Income", pretax_income),
        ("Income Tax Expense", income_tax),
        ("Net Income", net_income),
        ("Net Margin", net_margin),
    ]),
    ("Per Share", [
        ("Diluted EPS ($)", eps_diluted),
        ("Diluted Shares (M)", shares_diluted),
    ]),
])

# ── Balance Sheet ──
ws_bs = wb.create_sheet("Balance Sheet")
write_sheet(ws_bs, "Balance Sheet", "GoDaddy Inc. (GDDY) — Source: SEC EDGAR XBRL — All figures in $M", [
    ("Assets", [
        ("Cash & Equivalents", cash),
        ("Accounts Receivable", receivables),
        ("Total Current Assets", current_assets),
        ("Property & Equipment (net)", ppe),
        ("Goodwill", goodwill),
        ("Intangible Assets (net)", intangibles),
        ("Total Assets", total_assets),
    ]),
    ("Liabilities", [
        ("Deferred Revenue (Current)", deferred_rev_current),
        ("Total Current Liabilities", current_liabilities),
        ("Long-Term Debt", lt_debt),
        ("Deferred Revenue (Non-Current)", deferred_rev_noncurrent),
        ("Total Liabilities", total_liabilities),
    ]),
    ("Equity", [
        ("Stockholders' Equity (Deficit)", equity),
    ]),
])

# ── Cash Flow ──
ws_cf = wb.create_sheet("Cash Flow")

fcf = {}
fcf_margin = {}
for y in FISCAL_YEARS:
    o = ocf.get(y)
    c = capex.get(y)
    r = revenue.get(y)
    fcf[y] = round(o - c, 1) if (o is not None and c is not None) else None
    fcf_margin[y] = round(fcf[y] / r, 3) if (fcf[y] and r) else None

write_sheet(ws_cf, "Cash Flow Statement", "GoDaddy Inc. (GDDY) — Source: SEC EDGAR XBRL — All figures in $M", [
    ("Operating Activities", [
        ("Net Income", net_income),
        ("Depreciation & Amortization", da),
        ("Stock-Based Compensation", sbc),
        ("Cash from Operations", ocf),
    ]),
    ("Investing Activities", [
        ("Capital Expenditures", capex),
        ("Acquisitions", acquisitions),
        ("Cash from Investing", investing),
    ]),
    ("Financing Activities", [
        ("Share Repurchases", repurchases),
        ("Cash from Financing", financing),
    ]),
    ("Free Cash Flow", [
        ("FCF (OCF - CapEx)", fcf),
        ("FCF Margin", fcf_margin),
    ]),
])

# ── Save ──
output = "/Users/igill/replacing-nerd-jobs/private/gddy-statements/gddy-financials.xlsx"
wb.save(output)
print(f"\nSaved: {output}")
print("Sheets: Income Statement, Balance Sheet, Cash Flow")
print(f"Years: {', '.join(f'FY{y}' for y in FISCAL_YEARS)}")

# Print summary of what we got
print("\n── Data Coverage ──")
for label, data in [("Revenue", revenue), ("Net Income", net_income), ("Total Assets", total_assets), ("OCF", ocf)]:
    filled = [y for y in FISCAL_YEARS if data.get(y) is not None]
    print(f"  {label}: {', '.join(str(y) for y in filled) if filled else 'NONE'}")
