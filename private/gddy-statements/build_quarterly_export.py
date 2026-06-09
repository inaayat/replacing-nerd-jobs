#!/usr/bin/env python3
"""Export GoDaddy quarterly XBRL data from SEC into an Excel workbook.

Covers the last 3 fiscal years of quarterly data (Q1 2023 – most recent).
"""

import json
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Download fresh data first
import subprocess
DATA_PATH = pathlib.Path(__file__).parent / "sec_data.json"
subprocess.run([
    "curl", "-s",
    "-H", "User-Agent: inaayat research@inaayat.xyz",
    "https://data.sec.gov/api/xbrl/companyfacts/CIK0001609711.json",
    "-o", str(DATA_PATH)
], check=True)

with open(DATA_PATH) as f:
    raw = json.load(f)

facts_gaap = raw["facts"].get("us-gaap", {})
OUTPUT_PATH = pathlib.Path(__file__).parent / "gddy-quarterly-data.xlsx"

# Define quarters we want (last 3 years)
# GoDaddy FY = calendar year. We want Q1 2023 through most recent available.
QUARTERS = []
for year in [2023, 2024, 2025, 2026]:
    for q in [1, 2, 3, 4]:
        QUARTERS.append((year, q))

QUARTER_LABELS = [f"Q{q} {y}" for y, q in QUARTERS]


def quarter_start_end(year, quarter):
    """Return expected (start, end) dates for a quarter."""
    starts = {1: f"{year}-01-01", 2: f"{year}-04-01", 3: f"{year}-07-01", 4: f"{year}-10-01"}
    ends = {1: f"{year}-03-31", 2: f"{year}-06-30", 3: f"{year}-09-30", 4: f"{year}-12-31"}
    return starts[quarter], ends[quarter]


def extract_quarterly(concept_name, unit="USD"):
    """Extract single-quarter values (3-month spans) from 10-Q and 10-K filings."""
    if concept_name not in facts_gaap:
        return {}

    entries = facts_gaap[concept_name].get("units", {}).get(unit, [])
    results = {}

    for e in entries:
        form = e.get("form", "")
        if form not in ("10-Q", "10-K"):
            continue

        start = e.get("start", "")
        end = e.get("end", "")

        if not end:
            continue

        # For balance sheet items (no start date), use end date to place in quarter
        if not start:
            end_year = int(end[:4])
            end_month = int(end[5:7])
            end_day = int(end[8:10])
            # Map end date to quarter
            if end_month == 3 and end_day == 31:
                q_key = (end_year, 1)
            elif end_month == 6 and end_day == 30:
                q_key = (end_year, 2)
            elif end_month == 9 and end_day == 30:
                q_key = (end_year, 3)
            elif end_month == 12 and end_day == 31:
                q_key = (end_year, 4)
            else:
                continue
            if q_key in [(y, q) for y, q in QUARTERS]:
                results[q_key] = e["val"]
            continue

        # For duration items, only take single-quarter spans (~3 months)
        start_year = int(start[:4])
        start_month = int(start[5:7])
        end_year = int(end[:4])
        end_month = int(end[5:7])

        months_span = (end_year - start_year) * 12 + (end_month - start_month)
        if months_span < 2 or months_span > 4:
            continue

        # Determine which quarter this is
        if end_month == 3:
            q_key = (end_year, 1)
        elif end_month == 6:
            q_key = (end_year, 2)
        elif end_month == 9:
            q_key = (end_year, 3)
        elif end_month == 12:
            q_key = (end_year, 4)
        else:
            continue

        if q_key in [(y, q) for y, q in QUARTERS]:
            results[q_key] = e["val"]

    return results


# ── Categorize concepts ──
INCOME_KEYWORDS = [
    "Revenue", "Sales", "CostOf", "GrossProfit", "Operating",
    "SellingGeneral", "Research", "Marketing", "Administrative",
    "Depreciation", "Amortization", "Interest", "IncomeLoss",
    "IncomeTax", "NetIncome", "Earnings", "EarningsPerShare",
    "WeightedAverage", "Comprehensive", "OtherIncome", "Expense",
    "Restructuring", "Impairment"
]

BALANCE_KEYWORDS = [
    "Asset", "Liabilit", "Equity", "Cash", "Receivable", "Inventory",
    "Payable", "Debt", "Goodwill", "Intangible", "Property",
    "AccruedLiabilit", "Deferred", "Contract", "Lease", "Capital",
    "RetainedEarnings", "Treasury", "Stock", "Deficit"
]

CASHFLOW_KEYWORDS = [
    "CashProvided", "CashUsed", "NetCash", "Payment", "Proceed",
    "Repurchase", "Dividend", "Acquisition", "Purchase", "Issuance",
    "CapitalExpenditure", "FreeCash"
]


def categorize(name):
    for kw in CASHFLOW_KEYWORDS:
        if kw.lower() in name.lower():
            return "Cash Flow"
    for kw in BALANCE_KEYWORDS:
        if kw.lower() in name.lower():
            return "Balance Sheet"
    for kw in INCOME_KEYWORDS:
        if kw.lower() in name.lower():
            return "Income Statement"
    return "Other"


# ── Process all concepts ──
print("Extracting quarterly data...")

sheets_data = {
    "Income Statement": [],
    "Balance Sheet": [],
    "Cash Flow": [],
    "Other": [],
}

for concept_name, concept_data in sorted(facts_gaap.items()):
    label = concept_data.get("label", concept_name)

    # Try different units
    values = extract_quarterly(concept_name, "USD")
    unit_used = "USD"
    if not values:
        values = extract_quarterly(concept_name, "USD/shares")
        unit_used = "USD/shares"
    if not values:
        values = extract_quarterly(concept_name, "shares")
        unit_used = "shares"
    if not values:
        values = extract_quarterly(concept_name, "pure")
        unit_used = "ratio"

    if not values:
        continue

    # Only include if data exists in our target range (2023+)
    relevant = {k: v for k, v in values.items() if k[0] >= 2023}
    if len(relevant) < 2:
        continue

    category = categorize(concept_name)
    sheets_data[category].append({
        "concept": concept_name,
        "label": label,
        "unit": unit_used,
        "values": relevant,
    })


# ── Build Excel ──
wb = Workbook()

header_font = Font(name="Calibri", bold=True, size=10)
title_font = Font(name="Calibri", bold=True, size=13)
concept_font = Font(name="Calibri", size=9, color="595B4A")
alt_fill = PatternFill(start_color="F8F4E8", end_color="F8F4E8", fill_type="solid")
thin_border = Border(bottom=Side(style='thin', color='E0E0E0'))
q4_fill = PatternFill(start_color="E8E0C0", end_color="E8E0C0", fill_type="solid")

COLORS = {
    "Income Statement": "C9DAE1",
    "Balance Sheet": "D4E4C8",
    "Cash Flow": "F5D8E8",
    "Other": "F0E8C0",
}

# Filter quarters to only those with data
all_found_quarters = set()
for items in sheets_data.values():
    for item in items:
        all_found_quarters.update(item["values"].keys())

active_quarters = sorted([q for q in QUARTERS if q in all_found_quarters])
active_labels = [f"Q{q} {y}" for y, q in active_quarters]


def write_sheet(ws, title, items, color):
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A2'] = f'GoDaddy Inc. (GDDY) — {len(items)} line items — Quarterly — Source: SEC EDGAR XBRL'
    ws['A2'].font = Font(name="Calibri", size=9, italic=True, color="595B4A")

    headers = ["Line Item", "XBRL Concept", "Unit"] + active_labels
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = fill
        if col > 3:
            cell.alignment = Alignment(horizontal='right')
            # Highlight Q4 columns
            if "Q4" in h:
                cell.fill = PatternFill(start_color="D4C490", end_color="D4C490", fill_type="solid")

    for i, item in enumerate(items):
        row = 5 + i
        ws.cell(row=row, column=1, value=item["label"]).border = thin_border
        ws.cell(row=row, column=2, value=item["concept"]).font = concept_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3, value=item["unit"]).font = concept_font
        ws.cell(row=row, column=3).border = thin_border

        for j, qk in enumerate(active_quarters, 4):
            cell = ws.cell(row=row, column=j)
            val = item["values"].get(qk)
            if val is not None:
                if item["unit"] == "USD":
                    cell.value = val / 1_000_000
                    cell.number_format = '#,##0.0'
                elif item["unit"] == "shares":
                    cell.value = val / 1_000_000
                    cell.number_format = '#,##0.0'
                else:
                    cell.value = val
                    cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
            cell.border = thin_border
            # Shade Q4 columns
            if qk[1] == 4:
                cell.fill = q4_fill

        if i % 2 == 1:
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = alt_fill

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    for col in range(4, 4 + len(active_quarters)):
        ws.column_dimensions[get_column_letter(col)].width = 13

    ws.freeze_panes = "D5"


# Create sheets
first = True
for sheet_name in ["Income Statement", "Balance Sheet", "Cash Flow", "Other"]:
    items = sheets_data[sheet_name]
    if not items:
        continue
    if first:
        ws = wb.active
        ws.title = sheet_name
        first = False
    else:
        ws = wb.create_sheet(sheet_name)
    write_sheet(ws, f"{sheet_name} (Quarterly)", items, COLORS[sheet_name])
    print(f"  {sheet_name}: {len(items)} line items")

# README sheet
ws_sum = wb.create_sheet("README", 0)
ws_sum['A1'] = "GoDaddy Inc. (GDDY) — Quarterly SEC XBRL Financial Data"
ws_sum['A1'].font = title_font
ws_sum['A3'] = "Source"
ws_sum['A3'].font = header_font
ws_sum['B3'] = "SEC EDGAR XBRL API: data.sec.gov/api/xbrl/companyfacts/CIK0001609711.json"
ws_sum['A4'] = "Period"
ws_sum['A4'].font = header_font
ws_sum['B4'] = f"{active_labels[0]} – {active_labels[-1]}"
ws_sum['A5'] = "Granularity"
ws_sum['A5'].font = header_font
ws_sum['B5'] = "Single-quarter values (3-month periods) from 10-Q and 10-K filings"
ws_sum['A6'] = "Units"
ws_sum['A6'].font = header_font
ws_sum['B6'] = "USD values in $Millions, shares in Millions, ratios as-is"
ws_sum['A8'] = "Notes:"
ws_sum['A8'].font = header_font
ws_sum['A9'] = "• Q4 columns are shaded — Q4 is reported in the 10-K (not a separate 10-Q)"
ws_sum['A10'] = "• Q4 = Full Year (10-K) minus Q1+Q2+Q3 for some items; SEC may not tag Q4 standalone"
ws_sum['A11'] = "• Balance sheet items show point-in-time values at quarter-end"
ws_sum['A12'] = "• Use this for: rolling forecasts, seasonality analysis, quarterly variance analysis"
ws_sum.column_dimensions['A'].width = 40
ws_sum.column_dimensions['B'].width = 75

wb.save(OUTPUT_PATH)

# Cleanup
DATA_PATH.unlink(missing_ok=True)

total = sum(len(v) for v in sheets_data.values())
print(f"\nSaved: {OUTPUT_PATH}")
print(f"Total: {total} line items across {len(active_quarters)} quarters ({active_labels[0]} – {active_labels[-1]})")
