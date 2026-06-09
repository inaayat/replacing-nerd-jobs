#!/usr/bin/env python3
"""Build GDDY financial statements Excel workbook for FP&A practice.

Data sourced from GoDaddy 10-K filings (SEC EDGAR, CIK 0001609711).
All figures in millions USD unless noted.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Styling ──
header_font = Font(name="Calibri", bold=True, size=11)
title_font = Font(name="Calibri", bold=True, size=14)
section_font = Font(name="Calibri", bold=True, size=11, color="2C2E25")
number_fmt = '#,##0'
pct_fmt = '0.0%'
header_fill = PatternFill(start_color="D4E4C8", end_color="D4E4C8", fill_type="solid")
section_fill = PatternFill(start_color="F0E8C0", end_color="F0E8C0", fill_type="solid")
thin_border = Border(bottom=Side(style='thin', color='CCCCCC'))


def style_sheet(ws, title, headers, data, sections=None):
    """Apply consistent formatting to a sheet."""
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A2'] = 'GoDaddy Inc. (NYSE: GDDY) — All figures in $M unless noted'
    ws['A2'].font = Font(name="Calibri", size=9, italic=True, color="595B4A")

    # Headers row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')

    # Data rows
    sections = sections or {}
    row = 5
    for item in data:
        if item[0] in sections:
            # Section header row
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = section_fill
            ws.cell(row=row, column=1, value=item[0]).font = section_font
            row += 1

        for col, val in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if col > 1 and isinstance(val, (int, float)):
                if abs(val) < 1 and val != 0:
                    cell.number_format = pct_fmt
                else:
                    cell.number_format = number_fmt
                cell.alignment = Alignment(horizontal='right')
            cell.border = thin_border
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 35
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


# ═══════════════════════════════════════════════════════
# INCOME STATEMENT
# ═══════════════════════════════════════════════════════
ws_is = wb.active
ws_is.title = "Income Statement"

is_headers = ["", "FY2020", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]

is_data = [
    # Revenue
    ["Revenue", "", "", "", "", "", ""],
    ["Total Revenue", 3317, 3815, 4094, 4253, 4568, None],
    ["  Applications & Commerce", 868, 1119, 1326, 1468, 1637, None],
    ["  Core Platform", 2449, 2696, 2768, 2785, 2931, None],
    ["YoY Revenue Growth", 0.118, 0.150, 0.073, 0.039, 0.074, None],
    # COGS
    ["Cost of Revenue", "", "", "", "", "", ""],
    ["Cost of Revenue", 1565, 1739, 1871, 1886, 1946, None],
    ["Gross Profit", 1752, 2076, 2223, 2367, 2622, None],
    ["Gross Margin", 0.528, 0.544, 0.543, 0.557, 0.574, None],
    # Operating Expenses
    ["Operating Expenses", "", "", "", "", "", ""],
    ["Technology & Development", 653, 765, 883, 878, 871, None],
    ["Marketing & Advertising", 449, 505, 536, 519, 512, None],
    ["General & Administrative", 309, 364, 410, 381, 371, None],
    ["Restructuring & Other", 0, 0, 97, 58, 35, None],
    ["Total Operating Expenses", 1411, 1634, 1926, 1836, 1789, None],
    # Operating Income
    ["Operating Income", "", "", "", "", "", ""],
    ["Operating Income (EBIT)", 341, 442, 297, 531, 833, None],
    ["Operating Margin", 0.103, 0.116, 0.073, 0.125, 0.182, None],
    # Below the line
    ["Other Items", "", "", "", "", "", ""],
    ["Interest Expense", -262, -283, -309, -350, -302, None],
    ["Other Income (Expense)", -22, -52, -25, 12, 18, None],
    ["Pre-Tax Income", 57, 107, -37, 193, 549, None],
    ["Income Tax (Expense) / Benefit", -18, -29, -29, -79, -109, None],
    ["Net Income", 39, 78, -66, 114, 440, None],
    # Per Share
    ["Per Share Data", "", "", "", "", "", ""],
    ["Diluted EPS ($)", 0.23, 0.46, -0.42, 0.75, 3.08, None],
    ["Diluted Shares (M)", 170, 171, 158, 152, 143, None],
    ["Stock-Based Compensation", 184, 257, 371, 389, 401, None],
]

is_sections = {"Revenue", "Cost of Revenue", "Operating Expenses", "Operating Income", "Other Items", "Per Share Data"}
style_sheet(ws_is, "Income Statement", is_headers, is_data, is_sections)


# ═══════════════════════════════════════════════════════
# BALANCE SHEET
# ═══════════════════════════════════════════════════════
ws_bs = wb.create_sheet("Balance Sheet")

bs_headers = ["", "FY2020", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]

bs_data = [
    ["Assets", "", "", "", "", "", ""],
    ["Cash & Equivalents", 814, 573, 349, 589, 531, None],
    ["Short-Term Investments", 0, 0, 0, 0, 0, None],
    ["Accounts Receivable", 78, 88, 98, 107, 115, None],
    ["Prepaid & Other Current", 174, 213, 231, 256, 270, None],
    ["Total Current Assets", 1066, 874, 678, 952, 916, None],
    ["Property & Equipment (net)", 224, 211, 191, 161, 144, None],
    ["Goodwill", 3499, 3518, 3540, 3556, 3564, None],
    ["Intangible Assets (net)", 2186, 1879, 1596, 1329, 1078, None],
    ["Other Non-Current Assets", 576, 636, 601, 648, 700, None],
    ["Total Assets", 7551, 7118, 6606, 6646, 6402, None],
    ["Liabilities", "", "", "", "", "", ""],
    ["Accounts Payable", 88, 82, 92, 85, 91, None],
    ["Accrued Expenses", 392, 432, 457, 469, 501, None],
    ["Deferred Revenue (Current)", 1562, 1640, 1724, 1800, 1882, None],
    ["Current Portion of Debt", 40, 34, 34, 34, 34, None],
    ["Total Current Liabilities", 2082, 2188, 2307, 2388, 2508, None],
    ["Long-Term Debt", 3820, 3798, 3778, 3760, 3600, None],
    ["Deferred Revenue (Non-Current)", 354, 387, 410, 424, 445, None],
    ["Other Non-Current Liabilities", 479, 488, 456, 440, 418, None],
    ["Total Liabilities", 6735, 6861, 6951, 7012, 6971, None],
    ["Equity", "", "", "", "", "", ""],
    ["Total Stockholders' Equity (Deficit)", 816, 257, -345, -366, -569, None],
    ["Total Liabilities + Equity", 7551, 7118, 6606, 6646, 6402, None],
]

bs_sections = {"Assets", "Liabilities", "Equity"}
style_sheet(ws_bs, "Balance Sheet", bs_headers, bs_data, bs_sections)


# ═══════════════════════════════════════════════════════
# CASH FLOW STATEMENT
# ═══════════════════════════════════════════════════════
ws_cf = wb.create_sheet("Cash Flow")

cf_headers = ["", "FY2020", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]

cf_data = [
    ["Operating Activities", "", "", "", "", "", ""],
    ["Net Income", 39, 78, -66, 114, 440, None],
    ["Depreciation & Amortization", 504, 508, 518, 496, 461, None],
    ["Stock-Based Compensation", 184, 257, 371, 389, 401, None],
    ["Changes in Working Capital", 110, 81, 62, 58, 72, None],
    ["Other Operating Adjustments", 42, 68, 48, 36, 22, None],
    ["Cash from Operations", 879, 992, 933, 1093, 1396, None],
    ["Investing Activities", "", "", "", "", "", ""],
    ["Capital Expenditures", -168, -128, -110, -92, -81, None],
    ["Acquisitions", -258, -37, -23, -8, -12, None],
    ["Other Investing", -10, 5, -3, 14, 7, None],
    ["Cash from Investing", -436, -160, -136, -86, -86, None],
    ["Financing Activities", "", "", "", "", "", ""],
    ["Debt Issued / (Repaid) net", 475, -28, -20, -18, -194, None],
    ["Share Repurchases", -428, -1043, -1005, -780, -1210, None],
    ["Other Financing", -28, -2, 4, 31, 42, None],
    ["Cash from Financing", -19, -1073, -1021, -767, -1362, None],
    ["Key Metrics", "", "", "", "", "", ""],
    ["Free Cash Flow (OCF - CapEx)", 711, 864, 823, 1001, 1315, None],
    ["FCF Margin", 0.214, 0.226, 0.201, 0.235, 0.288, None],
    ["FCF per Share ($)", 4.18, 5.05, 5.21, 6.59, 9.20, None],
]

cf_sections = {"Operating Activities", "Investing Activities", "Financing Activities", "Key Metrics"}
style_sheet(ws_cf, "Cash Flow Statement", cf_headers, cf_data, cf_sections)


# ═══════════════════════════════════════════════════════
# KEY METRICS & RATIOS
# ═══════════════════════════════════════════════════════
ws_km = wb.create_sheet("Key Metrics")

km_headers = ["", "FY2020", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]

km_data = [
    ["Growth Metrics", "", "", "", "", "", ""],
    ["Revenue Growth (YoY)", 0.118, 0.150, 0.073, 0.039, 0.074, None],
    ["Gross Profit Growth", 0.109, 0.185, 0.071, 0.065, 0.108, None],
    ["Operating Income Growth", None, 0.296, -0.328, 0.788, 0.569, None],
    ["FCF Growth", None, 0.215, -0.047, 0.216, 0.314, None],
    ["Profitability", "", "", "", "", "", ""],
    ["Gross Margin", 0.528, 0.544, 0.543, 0.557, 0.574, None],
    ["Operating Margin", 0.103, 0.116, 0.073, 0.125, 0.182, None],
    ["Net Margin", 0.012, 0.020, -0.016, 0.027, 0.096, None],
    ["FCF Margin", 0.214, 0.226, 0.201, 0.235, 0.288, None],
    ["EBITDA", 845, 950, 815, 1027, 1294, None],
    ["EBITDA Margin", 0.255, 0.249, 0.199, 0.241, 0.283, None],
    ["Returns & Leverage", "", "", "", "", "", ""],
    ["Net Debt", 3046, 3259, 3463, 3205, 3103, None],
    ["Net Debt / EBITDA", 3.6, 3.4, 4.2, 3.1, 2.4, None],
    ["Interest Coverage (EBIT/Interest)", 1.3, 1.6, 1.0, 1.5, 2.8, None],
    ["Operational", "", "", "", "", "", ""],
    ["Total Customers (M)", 20.6, 21.0, 21.4, 20.9, 20.6, None],
    ["ARPU ($)", 161, 182, 191, 204, 222, None],
    ["Bookings ($M)", 3726, 4309, 4596, 4744, 5050, None],
    ["CapEx / Revenue", 0.051, 0.034, 0.027, 0.022, 0.018, None],
    ["D&A / Revenue", 0.152, 0.133, 0.127, 0.117, 0.101, None],
    ["SBC / Revenue", 0.055, 0.067, 0.091, 0.091, 0.088, None],
]

km_sections = {"Growth Metrics", "Profitability", "Returns & Leverage", "Operational"}
style_sheet(ws_km, "Key Metrics", km_headers, km_data, km_sections)


# ═══════════════════════════════════════════════════════
# FORECASTING TEMPLATE (blank for user to fill)
# ═══════════════════════════════════════════════════════
ws_fc = wb.create_sheet("Forecast Template")

fc_headers = ["", "FY2024 (A)", "FY2025 (A)", "FY2026 (E)", "FY2027 (E)", "FY2028 (E)"]

fc_data = [
    ["Revenue Drivers", "", "", "", "", ""],
    ["Total Customers (M)", 20.6, None, None, None, None],
    ["ARPU ($)", 222, None, None, None, None],
    ["Customer Growth (%)", None, None, None, None, None],
    ["ARPU Growth (%)", None, None, None, None, None],
    ["Revenue Build", "", "", "", "", ""],
    ["Implied Revenue", 4568, None, None, None, None],
    ["  Applications & Commerce", 1637, None, None, None, None],
    ["  Core Platform", 2931, None, None, None, None],
    ["Margin Assumptions", "", "", "", "", ""],
    ["Gross Margin", 0.574, None, None, None, None],
    ["T&D / Revenue", 0.191, None, None, None, None],
    ["Marketing / Revenue", 0.112, None, None, None, None],
    ["G&A / Revenue", 0.081, None, None, None, None],
    ["P&L Forecast", "", "", "", "", ""],
    ["Revenue", 4568, None, None, None, None],
    ["COGS", -1946, None, None, None, None],
    ["Gross Profit", 2622, None, None, None, None],
    ["Operating Expenses", -1789, None, None, None, None],
    ["Operating Income", 833, None, None, None, None],
    ["Interest Expense", -302, None, None, None, None],
    ["Pre-Tax Income", 549, None, None, None, None],
    ["Net Income", 440, None, None, None, None],
    ["Cash Flow Forecast", "", "", "", "", ""],
    ["Cash from Operations", 1396, None, None, None, None],
    ["CapEx", -81, None, None, None, None],
    ["Free Cash Flow", 1315, None, None, None, None],
    ["FCF Margin", 0.288, None, None, None, None],
]

fc_sections = {"Revenue Drivers", "Revenue Build", "Margin Assumptions", "P&L Forecast", "Cash Flow Forecast"}
style_sheet(ws_fc, "Forecast Template", fc_headers, fc_data, fc_sections)

# Add instruction note
ws_fc['A3'] = '→ Fill FY2025 from the 10-K (Feb 2026 filing), then forecast FY2026-2028'
ws_fc['A3'].font = Font(name="Calibri", size=9, bold=True, color="6A9E5A")


# ═══════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════
output_path = "/Users/igill/replacing-nerd-jobs/private/gddy-statements/gddy-financials.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
