#!/usr/bin/env python3
"""Export ALL GoDaddy XBRL data from SEC into a comprehensive Excel workbook.

Creates one sheet per financial statement category, with every available
line item organized for FP&A practice.
"""

import json
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_PATH = pathlib.Path(__file__).parent / "sec_data.json"
OUTPUT_PATH = pathlib.Path(__file__).parent / "gddy-full-xbrl-data.xlsx"

with open(DATA_PATH) as f:
    raw = json.load(f)

facts_gaap = raw["facts"].get("us-gaap", {})
facts_dei = raw["facts"].get("dei", {})

FISCAL_YEARS = list(range(2018, 2026))

# ── Extract helpers ──

def extract_annual_values(concept_data, unit="USD"):
    """Extract annual 10-K values keyed by fiscal year (using end-date)."""
    entries = concept_data.get("units", {}).get(unit, [])
    results = {}

    for e in entries:
        if e.get("form") != "10-K" or e.get("fp") != "FY":
            continue
        end = e.get("end", "")
        if not end:
            continue

        end_year = int(end[:4])
        end_month = int(end[5:7])
        fy = end_year if end_month >= 10 else end_year - 1

        # For items with start date, ensure ~12 month span (skip quarterly segments)
        start = e.get("start", "")
        if start:
            start_year = int(start[:4])
            start_month = int(start[5:7])
            months_span = (end_year - start_year) * 12 + (end_month - start_month)
            if months_span < 10:
                continue

        if fy in FISCAL_YEARS:
            results[fy] = e["val"]

    return results


# ── Categorize concepts ──

INCOME_STMT_KEYWORDS = [
    "Revenue", "Sales", "CostOf", "GrossProfit", "Operating",
    "SellingGeneral", "Research", "Marketing", "Administrative",
    "Depreciation", "Amortization", "Interest", "IncomeLoss",
    "IncomeTax", "NetIncome", "Earnings", "EarningsPerShare",
    "WeightedAverage", "Comprehensive", "OtherIncome", "Expense",
    "Restructuring", "Impairment"
]

BALANCE_SHEET_KEYWORDS = [
    "Asset", "Liabilit", "Equity", "Cash", "Receivable", "Inventory",
    "Payable", "Debt", "Goodwill", "Intangible", "Property",
    "AccruedLiabilit", "Deferred", "Contract", "Lease", "Capital",
    "RetainedEarnings", "Treasury", "Stock", "Deficit"
]

CASH_FLOW_KEYWORDS = [
    "CashProvided", "CashUsed", "NetCash", "Payment", "Proceed",
    "Repurchase", "Dividend", "Acquisition", "Purchase", "Issuance",
    "CapitalExpenditure", "FreeCash"
]

def categorize(concept_name):
    for kw in CASH_FLOW_KEYWORDS:
        if kw.lower() in concept_name.lower():
            return "Cash Flow"
    for kw in BALANCE_SHEET_KEYWORDS:
        if kw.lower() in concept_name.lower():
            return "Balance Sheet"
    for kw in INCOME_STMT_KEYWORDS:
        if kw.lower() in concept_name.lower():
            return "Income Statement"
    return "Other"


# ── Process all concepts ──
sheets_data = {
    "Income Statement": [],
    "Balance Sheet": [],
    "Cash Flow": [],
    "Other": [],
}

for concept_name, concept_data in sorted(facts_gaap.items()):
    label = concept_data.get("label", concept_name)
    description = concept_data.get("description", "")

    # Try USD first, then USD/shares, then shares
    values = extract_annual_values(concept_data, "USD")
    unit_used = "USD"
    if not values:
        values = extract_annual_values(concept_data, "USD/shares")
        unit_used = "USD/shares"
    if not values:
        values = extract_annual_values(concept_data, "shares")
        unit_used = "shares"
    if not values:
        values = extract_annual_values(concept_data, "pure")
        unit_used = "ratio"

    if not values:
        continue

    # Only include if at least 2 years of data
    if len(values) < 2:
        continue

    category = categorize(concept_name)
    sheets_data[category].append({
        "concept": concept_name,
        "label": label,
        "description": description,
        "unit": unit_used,
        "values": values,
    })

# Also add DEI facts (entity info)
dei_items = []
for concept_name, concept_data in sorted(facts_dei.items()):
    label = concept_data.get("label", concept_name)
    for unit_type in ["USD", "shares", "pure", "USD/shares"]:
        values = extract_annual_values(concept_data, unit_type)
        if values and len(values) >= 2:
            dei_items.append({
                "concept": concept_name,
                "label": label,
                "description": concept_data.get("description", ""),
                "unit": unit_type,
                "values": values,
            })
            break

# ── Build Excel ──
wb = Workbook()

header_font = Font(name="Calibri", bold=True, size=10)
title_font = Font(name="Calibri", bold=True, size=13)
concept_font = Font(name="Calibri", size=9, color="595B4A")
header_fill = PatternFill(start_color="D4E4C8", end_color="D4E4C8", fill_type="solid")
alt_fill = PatternFill(start_color="F8F4E8", end_color="F8F4E8", fill_type="solid")
thin_border = Border(bottom=Side(style='thin', color='E0E0E0'))

COLORS = {
    "Income Statement": "C9DAE1",
    "Balance Sheet": "D4E4C8",
    "Cash Flow": "F5D8E8",
    "Other": "F0E8C0",
    "Entity Info": "E8CCCC",
}


def write_data_sheet(ws, title, items, color):
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A2'] = f'GoDaddy Inc. (GDDY) — {len(items)} line items — Source: SEC EDGAR XBRL (CIK 0001609711)'
    ws['A2'].font = Font(name="Calibri", size=9, italic=True, color="595B4A")

    # Headers
    headers = ["Line Item", "XBRL Concept", "Unit"] + [f"FY{y}" for y in FISCAL_YEARS]
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = fill
        if col > 3:
            cell.alignment = Alignment(horizontal='right')

    # Data rows
    for i, item in enumerate(items):
        row = 5 + i
        ws.cell(row=row, column=1, value=item["label"]).border = thin_border
        ws.cell(row=row, column=2, value=item["concept"]).font = concept_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3, value=item["unit"]).font = concept_font
        ws.cell(row=row, column=3).border = thin_border

        for j, y in enumerate(FISCAL_YEARS, 4):
            cell = ws.cell(row=row, column=j)
            val = item["values"].get(y)
            if val is not None:
                if item["unit"] == "USD":
                    cell.value = val / 1_000_000
                    cell.number_format = '#,##0.0'
                elif item["unit"] == "shares":
                    cell.value = val / 1_000_000
                    cell.number_format = '#,##0.0'
                else:
                    cell.value = val
                    cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
            cell.border = thin_border

        # Alternate row shading
        if i % 2 == 1:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = alt_fill

    # Column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    for col in range(4, 4 + len(FISCAL_YEARS)):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Freeze panes
    ws.freeze_panes = "D5"


# Create sheets
first = True
for sheet_name in ["Income Statement", "Balance Sheet", "Cash Flow", "Other"]:
    items = sheets_data[sheet_name]
    if not items:
        continue
    if first:
        ws = wb.active
        ws.title = sheet_name
        first = False
    else:
        ws = wb.create_sheet(sheet_name)
    write_data_sheet(ws, sheet_name, items, COLORS[sheet_name])
    print(f"  {sheet_name}: {len(items)} line items")

# Entity info sheet
if dei_items:
    ws = wb.create_sheet("Entity Info")
    write_data_sheet(ws, "Entity Information (DEI)", dei_items, COLORS["Entity Info"])
    print(f"  Entity Info: {len(dei_items)} items")

# ── Summary sheet at the front ──
ws_sum = wb.create_sheet("README", 0)
ws_sum['A1'] = "GoDaddy Inc. (GDDY) — Complete SEC XBRL Financial Data"
ws_sum['A1'].font = title_font
ws_sum['A3'] = "Source"
ws_sum['A3'].font = header_font
ws_sum['B3'] = "SEC EDGAR XBRL API: data.sec.gov/api/xbrl/companyfacts/CIK0001609711.json"
ws_sum['A4'] = "CIK"
ws_sum['A4'].font = header_font
ws_sum['B4'] = "0001609711"
ws_sum['A5'] = "Period"
ws_sum['A5'].font = header_font
ws_sum['B5'] = f"FY{FISCAL_YEARS[0]} – FY{FISCAL_YEARS[-1]}"
ws_sum['A6'] = "Units"
ws_sum['A6'].font = header_font
ws_sum['B6'] = "USD values in $Millions, shares in Millions, ratios as-is"
ws_sum['A7'] = "Filing Type"
ws_sum['A7'].font = header_font
ws_sum['B7'] = "10-K (Annual Reports) only — full fiscal year data"

ws_sum['A9'] = "Sheets in this workbook:"
ws_sum['A9'].font = header_font
row = 10
for sheet_name, items in sheets_data.items():
    if items:
        ws_sum.cell(row=row, column=1, value=f"  {sheet_name}")
        ws_sum.cell(row=row, column=2, value=f"{len(items)} line items")
        row += 1
if dei_items:
    ws_sum.cell(row=row, column=1, value="  Entity Info")
    ws_sum.cell(row=row, column=2, value=f"{len(dei_items)} items")
    row += 1

ws_sum.cell(row=row + 1, column=1, value="Notes:").font = header_font
ws_sum.cell(row=row + 2, column=1, value="• Each row is one XBRL concept (line item) from the filing")
ws_sum.cell(row=row + 3, column=1, value="• 'XBRL Concept' column = the official taxonomy tag (useful for cross-company comparison)")
ws_sum.cell(row=row + 4, column=1, value="• Values use the most recent filing for each fiscal year (includes restatements)")
ws_sum.cell(row=row + 5, column=1, value="• Use this data to build forecasting models, variance analysis, and 3-statement models")

ws_sum.column_dimensions['A'].width = 40
ws_sum.column_dimensions['B'].width = 70

# Save
wb.save(OUTPUT_PATH)
total_items = sum(len(v) for v in sheets_data.values()) + len(dei_items)
print(f"\nSaved: {OUTPUT_PATH}")
print(f"Total: {total_items} line items across {len(FISCAL_YEARS)} fiscal years")
